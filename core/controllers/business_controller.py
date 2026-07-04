from __future__ import annotations

import io
import logging
import uuid

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.security import get_current_tenant_user
from config.database import get_db
from controllers.telegram_controller import prime_human_agent_cache
from dtos import (
    BusinessConfigResponse,
    BusinessConfigUpdateRequest,
    ProductCreateRequest,
    ProductResponse,
    ProductUpdateRequest,
)
from services import BusinessConfigService, ProductService
from services.order_service import OrderService
from services.product_service import FIELD_NAMES

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/business/me",
    tags=["business"],
    dependencies=[Depends(get_current_tenant_user)],
)


@router.get("/profile", response_model=BusinessConfigResponse)
def get_profile(db: Session = Depends(get_db)) -> BusinessConfigResponse:
    try:
        config = BusinessConfigService(db).get_config()
        return BusinessConfigResponse.model_validate(config)
    except Exception as exc:
        logger.error("business.get_profile failed: %s", exc)
        raise HTTPException(500, "Failed to retrieve profile")


@router.put("/profile", response_model=BusinessConfigResponse)
def update_profile(
    data: BusinessConfigUpdateRequest,
    db: Session = Depends(get_db),
) -> BusinessConfigResponse:
    try:
        config_service = BusinessConfigService(db)
        with db.begin():
            config = config_service.update_config(
                name=data.name,
                email=data.email,
                phone=data.phone,
                address=data.address,
                city=data.city,
                website=data.website,
                logo_url=data.logo_url,
                business_hours=data.business_hours,
                promotions_config=(
                    data.promotions_config.model_dump(mode="json")
                    if data.promotions_config is not None
                    else None
                ),
                best_sellers_config=(
                    data.best_sellers_config.model_dump(mode="json")
                    if data.best_sellers_config is not None
                    else None
                ),
                favorites_config=(
                    data.favorites_config.model_dump(mode="json")
                    if data.favorites_config is not None
                    else None
                ),
                estimated_attention_minutes=data.estimated_attention_minutes,
                human_agent_available=data.human_agent_available,
            )
        prime_human_agent_cache(config.human_agent_available)
        return BusinessConfigResponse.model_validate(config)
    except Exception as exc:
        logger.error("business.update_profile failed: %s", exc)
        raise HTTPException(500, "Failed to update profile")


@router.get("/attention-time")
def get_attention_time(db: Session = Depends(get_db)) -> dict[str, int | None]:
    try:
        config = BusinessConfigService(db).get_config()
        metrics = OrderService(db).get_attention_time_metrics()
        return {
            "estimated_attention_minutes": config.estimated_attention_minutes,
            **metrics,
        }
    except Exception as exc:
        logger.error("business.get_attention_time failed: %s", exc)
        raise HTTPException(500, "Failed to retrieve attention time metrics")


@router.get("/products", response_model=list[ProductResponse])
def list_products(
    category: str | None = None,
    available_only: bool = False,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
) -> list[ProductResponse]:
    try:
        products = ProductService(db).list_products(
            category=category,
            available_only=available_only,
            skip=skip,
            limit=limit,
        )
        return [ProductResponse.model_validate(product) for product in products]
    except Exception as exc:
        logger.error("business.list_products failed: %s", exc)
        raise HTTPException(500, "Failed to list products")


@router.post("/products", response_model=ProductResponse)
def create_product(
    data: ProductCreateRequest,
    db: Session = Depends(get_db),
) -> ProductResponse:
    try:
        with db.begin():
            product = ProductService(db).create_product(
                sku=data.sku,
                name=data.name,
                description=data.description,
                price=data.price,
                stock=data.stock,
                category=data.category,
                is_available=data.is_available,
                cost=data.cost,
                margin=data.margin,
                provider=data.provider,
                taxes=data.taxes,
                unit_of_measure=data.unit_of_measure,
                format=data.format,
            )
        return ProductResponse.model_validate(product)
    except Exception as exc:
        logger.error("business.create_product failed: %s", exc)
        raise HTTPException(500, "Failed to create product")


@router.put("/products/{product_id}", response_model=ProductResponse)
def update_product(
    product_id: str,
    data: ProductUpdateRequest,
    db: Session = Depends(get_db),
) -> ProductResponse:
    try:
        with db.begin():
            product = ProductService(db).update_product(
                product_id=uuid.UUID(product_id),
                sku=data.sku,
                name=data.name,
                description=data.description,
                price=data.price,
                stock=data.stock,
                category=data.category,
                is_available=data.is_available,
                cost=data.cost,
                margin=data.margin,
                provider=data.provider,
                taxes=data.taxes,
                unit_of_measure=data.unit_of_measure,
                format=data.format,
            )
        return ProductResponse.model_validate(product)
    except Exception as exc:
        logger.error("business.update_product failed: %s", exc)
        raise HTTPException(500, "Failed to update product")


@router.delete("/products/{product_id}")
def delete_product(product_id: str, db: Session = Depends(get_db)) -> dict:
    try:
        with db.begin():
            deleted = ProductService(db).delete_product(uuid.UUID(product_id))
        if not deleted:
            raise HTTPException(404, "Product not found")
        return {"status": "deleted", "id": product_id}
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("business.delete_product failed: %s", exc)
        raise HTTPException(500, "Failed to delete product")


@router.get("/products/export")
def export_products(db: Session = Depends(get_db)) -> StreamingResponse:
    try:
        buffer = ProductService(db).export_to_workbook()
        headers = {
            "Content-Disposition": 'attachment; filename="productos.xlsx"',
        }
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as exc:
        logger.error("business.export_products failed: %s", exc)
        raise HTTPException(500, "Failed to export products")


@router.get("/products/export/template")
def export_template(db: Session = Depends(get_db)) -> StreamingResponse:
    try:
        buffer = ProductService(db).export_template_workbook()
        headers = {
            "Content-Disposition": 'attachment; filename="plantilla_productos.xlsx"',
        }
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as exc:
        logger.error("business.export_template failed: %s", exc)
        raise HTTPException(500, "Failed to export template")


@router.post("/products/import")
def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> dict:
    try:
        content = file.file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        rows: list[dict[str, object]] = []
        for row in workbook.active.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            rows.append(
                {
                    FIELD_NAMES[index]: row[index] if index < len(row) else None
                    for index in range(len(FIELD_NAMES))
                }
            )
        result = ProductService(db).import_from_rows(rows)
        return {
            "status": "ok",
            "rows_processed": len(rows),
            **result,
        }
    except Exception as exc:
        logger.error("business.import_products failed: %s", exc)
        raise HTTPException(500, "Failed to import products")
