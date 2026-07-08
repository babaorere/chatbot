from __future__ import annotations

import io
import logging
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.security import get_admin_api_key
from config.database import get_db
from config.value_limits import (
    PAGINATION_LIMIT_MAX,
    PAGINATION_LIMIT_MIN,
    PAGINATION_SKIP_MAX,
    PAGINATION_SKIP_MIN,
)
from controllers.telegram_controller import (
    prime_business_config_cache,
    prime_human_agent_cache,
    refresh_catalog_cache_after_commit,
)
from services import (
    BusinessConfigService,
    KBService,
    ProductService,
    UserService,
    ConversationService,
)
from services.product_service import FIELD_NAMES
from dtos import (
    BusinessConfigUpdateRequest,
    BusinessConfigResponse,
    ProductCreateRequest,
    ProductUpdateRequest,
    ProductResponse,
    KBEntryCreateRequest,
    KBEntryUpdateRequest,
    KBEntryResponse,
    KBSearchRequest,
    KBSearchResponse,
    KBSearchResultItem,
)

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/business/config",
    tags=["business-config"],
    dependencies=[Depends(get_admin_api_key)],
)


# ── Profile ──────────────────────────────────────────────────────────────────


@router.get("/profile", response_model=BusinessConfigResponse)
def get_profile(
    db: Session = Depends(get_db),
) -> BusinessConfigResponse:
    """Recupera la configuración general del negocio usada por la aplicación."""
    try:
        config_service = BusinessConfigService(db)
        config = config_service.get_config()
        return BusinessConfigResponse.model_validate(config)
    except Exception as e:
        logger.error("get_profile failed: %s", e)
        raise HTTPException(500, "Failed to retrieve profile")


@router.put("/profile", response_model=BusinessConfigResponse)
def update_profile(
    data: BusinessConfigUpdateRequest,
    db: Session = Depends(get_db),
) -> BusinessConfigResponse:
    """Actualiza la configuración general del negocio dentro de una transacción."""
    try:
        config_service = BusinessConfigService(db)
        with db.begin():
            config = config_service.update_config(
                name=data.name,
                email=data.email,
                phone=data.phone,
                address=data.address,
                city=data.city,
                website=data.website,
                logo_url=data.logo_url,
                business_hours=data.business_hours,
                promotions_config=(
                    data.promotions_config.model_dump(mode="json")
                    if data.promotions_config is not None
                    else None
                ),
                best_sellers_config=(
                    data.best_sellers_config.model_dump(mode="json")
                    if data.best_sellers_config is not None
                    else None
                ),
                favorites_config=(
                    data.favorites_config.model_dump(mode="json")
                    if data.favorites_config is not None
                    else None
                ),
                estimated_attention_minutes=data.estimated_attention_minutes,
                human_agent_available=data.human_agent_available,
            )
        prime_business_config_cache(config)
        prime_human_agent_cache(config.human_agent_available)
        return BusinessConfigResponse.model_validate(config)
    except Exception as e:
        logger.error("update_profile failed: %s", e)
        raise HTTPException(500, "Failed to update profile")


# ── Products ─────────────────────────────────────────────────────────────────


@router.get("/products", response_model=list[ProductResponse])
def list_products(
    category: str | None = None,
    available_only: bool = False,
    skip: int = Query(default=0, ge=PAGINATION_SKIP_MIN, le=PAGINATION_SKIP_MAX),
    limit: int = Query(default=50, ge=PAGINATION_LIMIT_MIN, le=PAGINATION_LIMIT_MAX),
    db: Session = Depends(get_db),
) -> list[ProductResponse]:
    """Lista productos del catálogo con filtros administrativos y paginación."""
    try:
        product_svc = ProductService(db)
        products = product_svc.list_products(
            category=category,
            available_only=available_only,
            skip=skip,
            limit=limit,
        )
        return [ProductResponse.model_validate(p) for p in products]
    except Exception as e:
        logger.error("list_products failed: %s", e)
        raise HTTPException(500, "Failed to list products")


@router.post("/products", response_model=ProductResponse)
def create_product(
    data: ProductCreateRequest,
    db: Session = Depends(get_db),
) -> ProductResponse:
    """Crea un producto nuevo en el catálogo administrado por el negocio."""
    try:
        product_svc = ProductService(db)
        with db.begin():
            product = product_svc.create_product(
                sku=data.sku,
                name=data.name,
                description=data.description,
                price=data.price,
                stock=data.stock,
                category=data.category,
                is_available=data.is_available,
                cost=data.cost,
                margin=data.margin,
                provider=data.provider,
                taxes=data.taxes,
                unit_of_measure=data.unit_of_measure,
                format=data.format,
            )
        refresh_catalog_cache_after_commit("business_config_product_created")
        return ProductResponse.model_validate(product)
    except Exception as e:
        logger.error("create_product failed: %s", e)
        raise HTTPException(500, "Failed to create product")


@router.put("/products/{product_id}", response_model=ProductResponse)
def update_product(
    product_id: str,
    data: ProductUpdateRequest,
    db: Session = Depends(get_db),
) -> ProductResponse:
    """Actualiza un producto existente del catálogo usando su UUID."""
    try:
        product_svc = ProductService(db)
        with db.begin():
            product = product_svc.update_product(
                product_id=uuid.UUID(product_id),
                sku=data.sku,
                name=data.name,
                description=data.description,
                price=data.price,
                stock=data.stock,
                category=data.category,
                is_available=data.is_available,
                cost=data.cost,
                margin=data.margin,
                provider=data.provider,
                taxes=data.taxes,
                unit_of_measure=data.unit_of_measure,
                format=data.format,
            )
        refresh_catalog_cache_after_commit("business_config_product_updated")
        return ProductResponse.model_validate(product)
    except Exception as e:
        logger.error("update_product failed: %s", e)
        raise HTTPException(500, "Failed to update product")


@router.delete("/products/{product_id}")
def delete_product(
    product_id: str,
    db: Session = Depends(get_db),
) -> dict:
    """Elimina un producto del catálogo y devuelve el identificador eliminado."""
    try:
        product_svc = ProductService(db)
        with db.begin():
            deleted = product_svc.delete_product(uuid.UUID(product_id))
        if not deleted:
            raise HTTPException(404, "Product not found")
        refresh_catalog_cache_after_commit("business_config_product_deleted")
        return {"status": "deleted", "id": product_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("delete_product failed: %s", e)
        raise HTTPException(500, "Failed to delete product")


@router.get("/products/categories", response_model=list[str])
def get_product_categories(
    db: Session = Depends(get_db),
) -> list[str]:
    """Lista las categorías actualmente usadas por los productos del catálogo."""
    try:
        product_svc = ProductService(db)
        return product_svc.get_categories()
    except Exception as e:
        logger.error("get_product_categories failed: %s", e)
        raise HTTPException(500, "Failed to retrieve categories")


# ── Excel Export / Import ────────────────────────────────────────────────────


@router.get("/products/export")
def export_products(
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Exporta todos los productos actuales a un archivo Excel."""
    try:
        product_svc = ProductService(db)
        buf = product_svc.export_to_workbook()
        headers = {"Content-Disposition": 'attachment; filename="productos.xlsx"'}
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as e:
        logger.error("export_products failed: %s", e)
        raise HTTPException(500, "Failed to export products")


@router.get("/products/export/template")
def export_products_template(
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Exporta una plantilla Excel vacía y documentada para importación de productos."""
    try:
        product_svc = ProductService(db)
        buf = product_svc.export_template_workbook()
        headers = {
            "Content-Disposition": 'attachment; filename="plantilla_productos.xlsx"'
        }
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as e:
        logger.error("export_products_template failed: %s", e)
        raise HTTPException(500, "Failed to export template")


@router.post("/products/import")
def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> dict:
    """
    Importa productos desde un archivo Excel (.xlsx).
    Estrategia de colisión: UPSERT por SKU.
    Si el SKU ya existe → se actualizan todos los campos.
    Si no existe → se crea un producto nuevo.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx")
    try:
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        # Primera fila = cabeceras (las ignoramos, usamos FIELD_NAMES por posición)
        rows: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Saltar filas completamente en blanco
            if all(cell is None for cell in row):
                continue
            row_dict = {
                FIELD_NAMES[i]: row[i] if i < len(row) else None
                for i in range(len(FIELD_NAMES))
            }
            rows.append(row_dict)

        product_svc = ProductService(db)
        with db.begin():
            summary = product_svc.import_from_rows(rows)
        refresh_catalog_cache_after_commit("business_config_products_imported")

        return {
            "status": "ok",
            "rows_processed": len(rows),
            **summary,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("import_products failed: %s", e)
        raise HTTPException(500, "Failed to import products")


# ── Knowledge Base ───────────────────────────────────────────────────────────


@router.get("/kb", response_model=list[KBEntryResponse])
def list_kb_entries(
    category: str | None = None,
    active_only: bool = True,
    skip: int = Query(default=0, ge=PAGINATION_SKIP_MIN, le=PAGINATION_SKIP_MAX),
    limit: int = Query(default=50, ge=PAGINATION_LIMIT_MIN, le=PAGINATION_LIMIT_MAX),
    db: Session = Depends(get_db),
) -> list[KBEntryResponse]:
    """Lista entradas de knowledge base con filtros administrativos y paginación."""
    try:
        kb_svc = KBService(db)
        entries = kb_svc.list_entries(
            category=category,
            active_only=active_only,
            skip=skip,
            limit=limit,
        )
        return [KBEntryResponse.model_validate(e) for e in entries]
    except Exception as e:
        logger.error("list_kb_entries failed: %s", e)
        raise HTTPException(500, "Failed to list KB entries")


@router.post("/kb", response_model=KBEntryResponse)
async def create_kb_entry(
    data: KBEntryCreateRequest,
    db: Session = Depends(get_db),
) -> KBEntryResponse:
    """Crea una nueva entrada en la knowledge base y genera sus artefactos asociados."""
    try:
        kb_svc = KBService(db)
        entry = await kb_svc.create_entry(
            category=data.category,
            title=data.title,
            content=data.content,
        )
        return KBEntryResponse.model_validate(entry)
    except Exception as e:
        logger.error("create_kb_entry failed: %s", e)
        raise HTTPException(500, "Failed to create KB entry")


@router.put("/kb/{entry_id}", response_model=KBEntryResponse)
async def update_kb_entry(
    entry_id: str,
    data: KBEntryUpdateRequest,
    db: Session = Depends(get_db),
) -> KBEntryResponse:
    """Actualiza una entrada existente de la knowledge base usando su UUID."""
    try:
        kb_svc = KBService(db)
        entry = await kb_svc.update_entry(
            entry_id=uuid.UUID(entry_id),
            category=data.category,
            title=data.title,
            content=data.content,
            is_active=data.is_active,
        )
        return KBEntryResponse.model_validate(entry)
    except Exception as e:
        logger.error("update_kb_entry failed: %s", e)
        raise HTTPException(500, "Failed to update KB entry")


@router.delete("/kb/{entry_id}")
def delete_kb_entry(
    entry_id: str,
    db: Session = Depends(get_db),
) -> dict:
    """Elimina una entrada de knowledge base y devuelve el identificador eliminado."""
    try:
        kb_svc = KBService(db)
        with db.begin():
            deleted = kb_svc.delete_entry(uuid.UUID(entry_id))
        if not deleted:
            raise HTTPException(404, "KB entry not found")
        return {"status": "deleted", "id": entry_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("delete_kb_entry failed: %s", e)
        raise HTTPException(500, "Failed to delete KB entry")


@router.post("/kb/search", response_model=KBSearchResponse)
async def search_kb(
    data: KBSearchRequest,
    db: Session = Depends(get_db),
) -> KBSearchResponse:
    """Ejecuta una búsqueda administrativa sobre la knowledge base indexada."""
    try:
        kb_svc = KBService(db)
        results = await kb_svc.search(
            query=data.query, top_k=data.top_k, category=data.category
        )
        return KBSearchResponse(
            query=data.query,
            results=[KBSearchResultItem(**r) for r in results],
            count=len(results),
        )
    except Exception as e:
        logger.error("search_kb failed: %s", e)
        raise HTTPException(500, "Failed to search KB")


@router.get("/kb/categories", response_model=list[str])
def get_kb_categories(
    db: Session = Depends(get_db),
) -> list[str]:
    """Lista las categorías disponibles en la knowledge base."""
    try:
        kb_svc = KBService(db)
        return kb_svc.get_categories()
    except Exception as e:
        logger.error("get_kb_categories failed: %s", e)
        raise HTTPException(500, "Failed to retrieve KB categories")


# ── Users & Conversations (read-only) ────────────────────────────────────────


@router.get("/users/count")
def get_user_count(
    db: Session = Depends(get_db),
) -> dict:
    """Devuelve el conteo total de usuarios persistidos."""
    try:
        user_svc = UserService(db)
        count = user_svc.repo.count()
        return {"count": count}
    except Exception as e:
        logger.error("get_user_count failed: %s", e)
        raise HTTPException(500, "Failed to get user count")


@router.get("/conversations/count")
def get_conversation_count(
    db: Session = Depends(get_db),
) -> dict:
    """Devuelve el conteo total de conversaciones persistidas."""
    try:
        conv_svc = ConversationService(db)
        count = conv_svc.repo.count()
        return {"count": count}
    except Exception as e:
        logger.error("get_conversation_count failed: %s", e)
        raise HTTPException(500, "Failed to get conversation count")
