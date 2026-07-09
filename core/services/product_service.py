from __future__ import annotations

import io
import logging
import math
import uuid
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from models.category import Category
from models.product import Product
from repositories.product_repository import ProductRepository
from services.category_service import slugify
from config.value_limits import (
    PRODUCT_MARGIN_MAX,
    PRODUCT_MARGIN_MIN,
    PRODUCT_MONEY_MAX,
    PRODUCT_MONEY_MIN,
    PRODUCT_STOCK_MAX,
    PRODUCT_STOCK_MIN,
    PRODUCT_TAX_MAX,
    PRODUCT_TAX_MIN,
    ensure_int_range,
    ensure_optional_float_range,
)

logger = logging.getLogger(__name__)

# ── Columnas del Excel (orden canónico) ───────────────────────────────────────

EXCEL_COLUMNS: list[tuple[str, str]] = [
    ("sku", "SKU"),
    ("name", "Nombre *"),
    ("description", "Descripción"),
    ("price", "Precio"),
    ("cost", "Costo"),
    ("stock", "Stock"),
    ("format", "Formato (ej: 500cc, caja x12, unidad)"),
    ("unit_of_measure", "Unidad de medida (ej: un, kg, lt)"),
    ("category", "Categoría"),
    ("provider", "Proveedor"),
    ("taxes", "IVA (0.0 – 1.0, ej: 0.19)"),
    ("is_available", "Disponible (TRUE/FALSE)"),
    ("margin", "Margen (%)"),
]

FIELD_NAMES = [col[0] for col in EXCEL_COLUMNS]
HEADER_LABELS = [col[1] for col in EXCEL_COLUMNS]


class ProductService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.repo = ProductRepository(db)

    # ── Helpers internos ─────────────────────────────────────────────────────

    def _ensure_category_exists(self, category_name: str | None) -> None:
        if not category_name:
            return
        exists = self.db.query(Category).filter(Category.name == category_name).first()
        if not exists:
            slug = slugify(category_name)
            existing_slug = (
                self.db.query(Category).filter(Category.slug == slug).first()
            )
            if existing_slug:
                slug = f"{slug}-auto"
            cat = Category(name=category_name, slug=slug, is_system=False)
            self.db.add(cat)
            self.db.flush()

    def _row_to_float(self, value: object) -> float | None:
        if value is None or str(value).strip() == "":
            return None
        try:
            parsed = float(str(value).strip())
            if not math.isfinite(parsed):
                return None
            return parsed
        except ValueError:
            return None

    def _row_to_int(self, value: object) -> int:
        if value is None or str(value).strip() == "":
            return 0
        try:
            parsed = float(str(value).strip())
            if not math.isfinite(parsed):
                return 0
            return int(parsed)
        except (OverflowError, ValueError):
            return 0

    def _row_to_bool(self, value: object) -> bool:
        if value is None:
            return True
        return str(value).strip().upper() not in {"FALSE", "0", "NO", "FALSO", "F"}

    def _row_to_str(self, value: object) -> str | None:
        if value is None:
            return None
        stripped = str(value).strip()
        return stripped if stripped else None

    # ── CRUD ─────────────────────────────────────────────────────────────────

    def _validate_product_limits(
        self,
        *,
        price: float | None = None,
        stock: int | None = None,
        cost: float | None = None,
        margin: float | None = None,
        taxes: float | None = None,
    ) -> None:
        ensure_optional_float_range(
            price,
            name="Precio",
            min_value=PRODUCT_MONEY_MIN,
            max_value=PRODUCT_MONEY_MAX,
        )
        if stock is not None:
            ensure_int_range(
                stock,
                name="Stock",
                min_value=PRODUCT_STOCK_MIN,
                max_value=PRODUCT_STOCK_MAX,
            )
        ensure_optional_float_range(
            cost,
            name="Costo",
            min_value=PRODUCT_MONEY_MIN,
            max_value=PRODUCT_MONEY_MAX,
        )
        ensure_optional_float_range(
            margin,
            name="Margen",
            min_value=PRODUCT_MARGIN_MIN,
            max_value=PRODUCT_MARGIN_MAX,
        )
        ensure_optional_float_range(
            taxes,
            name="IVA",
            min_value=PRODUCT_TAX_MIN,
            max_value=PRODUCT_TAX_MAX,
        )

    def list_products(
        self,
        category: str | None = None,
        available_only: bool = False,
        skip: int = 0,
        limit: int = 50,
    ) -> list[Product]:
        try:
            return self.repo.find_all(
                category=category,
                available_only=available_only,
                skip=skip,
                limit=limit,
            )
        except Exception as e:
            logger.error("ProductService.list_products failed: %s", e)
            raise

    def get_product(self, product_id: uuid.UUID) -> Product | None:
        try:
            return self.repo.find_by_id(product_id)
        except Exception as e:
            logger.error(
                "ProductService.get_product failed [id=%s]: %s",
                product_id,
                e,
            )
            raise

    def create_product(
        self,
        name: str,
        sku: str | None = None,
        description: str | None = None,
        price: float | None = None,
        stock: int = 0,
        category: str | None = None,
        is_available: bool = True,
        cost: float | None = None,
        margin: float | None = None,
        provider: str | None = None,
        taxes: float | None = 0.19,
        unit_of_measure: str | None = "un",
        format: str | None = None,
    ) -> Product:
        try:
            self._validate_product_limits(
                price=price,
                stock=stock,
                cost=cost,
                margin=margin,
                taxes=taxes,
            )
            self._ensure_category_exists(category)
            product = Product(
                sku=sku,
                name=name,
                description=description,
                price=price,
                stock=stock,
                category=category,
                is_available=is_available,
                cost=cost,
                margin=margin,
                provider=provider,
                taxes=taxes,
                unit_of_measure=unit_of_measure,
                format=format,
            )
            return self.repo.save(product)
        except Exception as e:
            logger.error(
                "ProductService.create_product failed [name=%s]: %s",
                name,
                e,
            )
            raise

    def update_product(
        self,
        product_id: uuid.UUID,
        sku: str | None = None,
        name: str | None = None,
        description: str | None = None,
        price: float | None = None,
        stock: int | None = None,
        category: str | None = None,
        is_available: bool | None = None,
        cost: float | None = None,
        margin: float | None = None,
        provider: str | None = None,
        taxes: float | None = None,
        unit_of_measure: str | None = None,
        format: str | None = None,
    ) -> Product:
        try:
            self._validate_product_limits(
                price=price,
                stock=stock,
                cost=cost,
                margin=margin,
                taxes=taxes,
            )
            product = self.repo.find_by_id(product_id)
            if not product:
                raise ValueError(f"Product {product_id} not found")

            if sku is not None:
                product.sku = sku
            if name is not None:
                product.name = name
            if description is not None:
                product.description = description
            if price is not None:
                product.price = price
            if stock is not None:
                product.stock = stock
            if category is not None:
                self._ensure_category_exists(category)
                product.category = category
            if is_available is not None:
                product.is_available = is_available
            if cost is not None:
                product.cost = cost
            if margin is not None:
                product.margin = margin
            if provider is not None:
                product.provider = provider
            if taxes is not None:
                product.taxes = taxes
            if unit_of_measure is not None:
                product.unit_of_measure = unit_of_measure
            if format is not None:
                product.format = format

            self.db.flush()
            self.db.refresh(product)
            return product
        except Exception as e:
            logger.error(
                "ProductService.update_product failed [id=%s]: %s",
                product_id,
                e,
            )
            raise

    def delete_product(self, product_id: uuid.UUID) -> bool:
        try:
            product = self.repo.find_by_id(product_id)
            if not product:
                return False
            self.db.delete(product)
            self.db.flush()
            return True
        except Exception as e:
            logger.error(
                "ProductService.delete_product failed [id=%s]: %s",
                product_id,
                e,
            )
            raise

    def search(self, query: str, limit: int = 20) -> list[Product]:
        try:
            return self.repo.search_by_name(query, limit=limit)
        except Exception as e:
            logger.error(
                "ProductService.search failed [query=%s]: %s",
                query,
                e,
            )
            raise

    def get_categories(self) -> list[str]:
        try:
            return self.repo.get_categories()
        except Exception as e:
            logger.error("ProductService.get_categories failed: %s", e)
            raise

    def count(self, category: str | None = None, available_only: bool = False) -> int:
        try:
            return self.repo.count_all(category=category, available_only=available_only)
        except Exception as e:
            logger.error("ProductService.count failed: %s", e)
            raise

    # ── Importación / Exportación Excel ──────────────────────────────────────

    def upsert_by_sku(
        self,
        row: dict[str, object],
    ) -> tuple[Product, bool]:
        """
        Crea o actualiza un producto basado en su SKU.
        Si el SKU ya existe → actualiza todos los campos presentes.
        Si no existe → crea uno nuevo.
        Retorna (producto, created: bool).
        """
        sku = self._row_to_str(row.get("sku"))
        name = self._row_to_str(row.get("name"))
        if not name:
            raise ValueError("El campo 'Nombre' es obligatorio.")

        category = self._row_to_str(row.get("category"))
        price = self._row_to_float(row.get("price"))
        cost = self._row_to_float(row.get("cost"))
        stock = self._row_to_int(row.get("stock"))
        taxes = self._row_to_float(row.get("taxes"))
        margin = self._row_to_float(row.get("margin"))
        self._validate_product_limits(
            price=price,
            stock=stock,
            cost=cost,
            margin=margin,
            taxes=taxes,
        )
        self._ensure_category_exists(category)

        existing: Product | None = None
        if sku:
            existing = self.repo.find_by_sku(sku)

        if existing:
            existing.name = name
            existing.description = self._row_to_str(row.get("description"))
            existing.price = price
            existing.cost = cost
            existing.stock = stock
            existing.format = self._row_to_str(row.get("format"))
            existing.unit_of_measure = (
                self._row_to_str(row.get("unit_of_measure")) or "un"
            )
            existing.category = category
            existing.provider = self._row_to_str(row.get("provider"))
            existing.taxes = taxes
            existing.is_available = self._row_to_bool(row.get("is_available"))
            existing.margin = margin
            self.db.flush()
            self.db.refresh(existing)
            return existing, False

        product = Product(
            sku=sku,
            name=name,
            description=self._row_to_str(row.get("description")),
            price=price,
            cost=cost,
            stock=stock,
            format=self._row_to_str(row.get("format")),
            unit_of_measure=self._row_to_str(row.get("unit_of_measure")) or "un",
            category=category,
            provider=self._row_to_str(row.get("provider")),
            taxes=taxes,
            is_available=self._row_to_bool(row.get("is_available")),
            margin=margin,
        )
        return self.repo.save(product), True

    def import_from_rows(self, rows: list[dict[str, object]]) -> dict[str, int]:
        """
        Importa una lista de filas (dicts con claves = FIELD_NAMES).
        Estrategia de colisión: UPSERT por SKU.
        Retorna un resumen: {"created": n, "updated": n, "errors": 0}.
        """
        created = 0
        updated = 0
        try:
            for index, row in enumerate(rows, start=2):
                try:
                    _, was_created = self.upsert_by_sku(row)
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                except Exception as row_exc:
                    logger.error("import_from_rows error on row %d: %s", index, row_exc)
                    raise ValueError(f"Fila {index}: {row_exc}") from row_exc
            return {"created": created, "updated": updated, "errors": 0}
        except Exception as exc:
            logger.error("import_from_rows failed: %s", exc)
            raise

    def load_import_rows_from_workbook_bytes(
        self,
        workbook_bytes: bytes,
    ) -> list[dict[str, object]]:
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
        worksheet = workbook.active
        rows: list[dict[str, object]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            row_dict = {
                FIELD_NAMES[index]: row[index] if index < len(row) else None
                for index in range(len(FIELD_NAMES))
            }
            rows.append(row_dict)
        return rows

    def import_from_workbook_bytes(self, workbook_bytes: bytes) -> dict[str, int]:
        rows = self.load_import_rows_from_workbook_bytes(workbook_bytes)
        summary = self.import_from_rows(rows)
        return {"rows_processed": len(rows), **summary}

    def export_to_workbook(self) -> io.BytesIO:
        """Exporta todos los productos al formato Excel (.xlsx)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center")

        for col_index, label in enumerate(HEADER_LABELS, start=1):
            cell = ws.cell(row=1, column=col_index, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        products = self.repo.find_all(skip=0, limit=10000)
        for row_index, product in enumerate(products, start=2):
            price = (
                float(product.price)
                if isinstance(product.price, Decimal)
                else product.price
            )
            cost = (
                float(product.cost)
                if isinstance(product.cost, Decimal)
                else product.cost
            )
            margin = (
                float(product.margin)
                if isinstance(product.margin, Decimal)
                else product.margin
            )
            taxes = (
                float(product.taxes)
                if isinstance(product.taxes, Decimal)
                else product.taxes
            )

            row_values = [
                product.sku,
                product.name,
                product.description,
                price,
                cost,
                product.stock,
                product.format,
                product.unit_of_measure,
                product.category,
                product.provider,
                taxes,
                product.is_available,
                margin,
            ]
            for col_index, value in enumerate(row_values, start=1):
                ws.cell(row=row_index, column=col_index, value=value)

        for col_index in range(1, len(HEADER_LABELS) + 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_index).column_letter
            ].width = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def export_template_workbook(self) -> io.BytesIO:
        """
        Exporta una hoja Excel vacía con cabeceras documentadas y
        una fila de ejemplo para orientar al usuario.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Productos"

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        example_fill = PatternFill("solid", fgColor="D9E1F2")
        center_align = Alignment(horizontal="center", vertical="center")

        for col_index, label in enumerate(HEADER_LABELS, start=1):
            cell = ws.cell(row=1, column=col_index, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        example_values = [
            "PROD-001",
            "Pisco Control 35°",
            "Botella de pisco 750ml grado 35",
            4990,
            3500,
            24,
            "750ml",
            "un",
            "Piscos",
            "CCU",
            0.19,
            True,
            0.30,
        ]
        for col_index, value in enumerate(example_values, start=1):
            cell = ws.cell(row=2, column=col_index, value=value)
            cell.fill = example_fill
            cell.alignment = center_align

        for col_index in range(1, len(HEADER_LABELS) + 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_index).column_letter
            ].width = 26

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
